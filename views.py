import json, math, secrets, io
from datetime import datetime, time
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, FileResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from django.views.decorators.http import require_POST
from .forms import LeaveRequestForm, OfficialDutyForm, TeacherImportForm
from .models import Attendance, LeaveRequest, OfficialDuty, DailyQR

def haversine_m(lat1, lon1, lat2, lon2):
    radius = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2-lat1)
    dl = math.radians(lon2-lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return radius * (2 * math.atan2(math.sqrt(a), math.sqrt(1-a)))

def allowed_times(today):
    # Monday=0 ... Sunday=6
    if today.weekday() == 4:
        return time(7,10), time(11,40)
    return time(7,10), time(13,0)

@login_required
def dashboard(request):
    today = timezone.localdate()
    record = Attendance.objects.filter(user=request.user, date=today).first()
    month_records = Attendance.objects.filter(
        user=request.user, date__year=today.year, date__month=today.month
    )
    context = {
        "record": record,
        "today": today,
        "month_count": month_records.count(),
        "late_count": month_records.filter(status="LEWAT").count(),
        "leave_pending": LeaveRequest.objects.filter(user=request.user, status="MENUNGGU").count(),
        "duty_pending": OfficialDuty.objects.filter(user=request.user, status="MENUNGGU").count(),
    }
    return render(request, "attendance/dashboard.html", context)

@login_required
@require_POST
def record_attendance(request, action):
    if action not in {"masuk", "keluar"}:
        return JsonResponse({"ok": False, "message": "Tindakan tidak sah."}, status=400)
    try:
        lat = float(request.POST.get("latitude", ""))
        lng = float(request.POST.get("longitude", ""))
    except ValueError:
        return JsonResponse({"ok": False, "message": "Lokasi GPS tidak sah."}, status=400)

    distance = haversine_m(
        lat, lng, settings.SCHOOL_LATITUDE, settings.SCHOOL_LONGITUDE
    )
    if distance > settings.SCHOOL_RADIUS_METERS:
        return JsonResponse({
            "ok": False,
            "message": f"Anda berada {distance:.1f} m dari sekolah. Had maksimum ialah {settings.SCHOOL_RADIUS_METERS} m."
        }, status=403)

    today = timezone.localdate()
    now = timezone.now()
    record, _ = Attendance.objects.get_or_create(user=request.user, date=today)
    selfie = request.FILES.get("selfie")
    check_in_limit, _ = allowed_times(today)

    if action == "masuk":
        if record.check_in:
            return JsonResponse({"ok": False, "message": "Rekod masuk sudah dibuat."}, status=400)
        record.check_in = now
        record.check_in_lat = lat
        record.check_in_lng = lng
        record.distance_in_m = distance
        record.selfie_in = selfie
        record.status = "LEWAT" if timezone.localtime(now).time() > check_in_limit else "HADIR"
        record.save()
        return JsonResponse({"ok": True, "message": f"Rekod masuk berjaya. Jarak {distance:.1f} m."})

    if not record.check_in:
        return JsonResponse({"ok": False, "message": "Sila rekod masuk terlebih dahulu."}, status=400)
    if record.check_out:
        return JsonResponse({"ok": False, "message": "Rekod keluar sudah dibuat."}, status=400)

    record.check_out = now
    record.check_out_lat = lat
    record.check_out_lng = lng
    record.distance_out_m = distance
    record.selfie_out = selfie
    record.save()
    return JsonResponse({"ok": True, "message": f"Rekod keluar berjaya. Jarak {distance:.1f} m."})

@login_required
def leave_page(request):
    if request.method == "POST":
        form = LeaveRequestForm(request.POST, request.FILES)
        if form.is_valid():
            item = form.save(commit=False)
            item.user = request.user
            item.save()
            messages.success(request, "Permohonan cuti berjaya dihantar.")
            return redirect("leave_page")
    else:
        form = LeaveRequestForm()
    items = LeaveRequest.objects.filter(user=request.user)
    return render(request, "attendance/leave.html", {"form": form, "items": items})

@login_required
def official_duty_page(request):
    if request.method == "POST":
        form = OfficialDutyForm(request.POST, request.FILES)
        if form.is_valid():
            item = form.save(commit=False)
            item.user = request.user
            item.save()
            messages.success(request, "Permohonan tugas rasmi berjaya dihantar.")
            return redirect("official_duty_page")
    else:
        form = OfficialDutyForm()
    items = OfficialDuty.objects.filter(user=request.user)
    return render(request, "attendance/official_duty.html", {"form": form, "items": items})

@login_required
def report_page(request):
    records = Attendance.objects.filter(user=request.user)[:100]
    return render(request, "attendance/report.html", {"records": records})

def manifest(request):
    data = {
        "name": "Kehadiran Guru SK Ulu Ansuan",
        "short_name": "Kehadiran",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#e5e7eb",
        "theme_color": "#4b5563",
        "icons": [{
            "src": "/static/attendance/icons/logo_sekolah.jpg",
            "sizes": "512x512",
            "type": "image/jpeg"
        }]
    }
    return JsonResponse(data)

def service_worker(request):
    script = '''
const CACHE = "kehadiran-v4";
const ASSETS = ["/", "/login/", "/static/attendance/css/style.css", "/static/attendance/js/app.js"];
self.addEventListener("install", e => e.waitUntil(caches.open(CACHE).then(c => c.addAll(ASSETS))));
self.addEventListener("fetch", e => e.respondWith(fetch(e.request).catch(() => caches.match(e.request))));
'''
    return HttpResponse(script, content_type="application/javascript")


@login_required
def export_excel(request):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Kehadiran"
    ws.append(["Tarikh", "Nama", "Masuk", "Keluar", "Status", "Jarak Masuk (m)", "Jarak Keluar (m)"])
    qs = Attendance.objects.filter(user=request.user).order_by("-date")
    for r in qs:
        ws.append([
            str(r.date),
            r.user.get_full_name() or r.user.username,
            timezone.localtime(r.check_in).strftime("%H:%M:%S") if r.check_in else "",
            timezone.localtime(r.check_out).strftime("%H:%M:%S") if r.check_out else "",
            r.get_status_display(),
            round(r.distance_in_m or 0, 1),
            round(r.distance_out_m or 0, 1),
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="laporan_kehadiran.xlsx"'
    return response

@login_required
def export_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, height - 45, "Laporan Kehadiran Guru SK Ulu Ansuan")
    c.setFont("Helvetica", 9)
    y = height - 75
    c.drawString(40, y, "Tarikh")
    c.drawString(110, y, "Masuk")
    c.drawString(165, y, "Keluar")
    c.drawString(220, y, "Status")
    c.drawString(300, y, "Jarak Masuk")
    y -= 18
    for r in Attendance.objects.filter(user=request.user).order_by("-date")[:150]:
        if y < 50:
            c.showPage()
            c.setFont("Helvetica", 9)
            y = height - 50
        c.drawString(40, y, str(r.date))
        c.drawString(110, y, timezone.localtime(r.check_in).strftime("%H:%M") if r.check_in else "-")
        c.drawString(165, y, timezone.localtime(r.check_out).strftime("%H:%M") if r.check_out else "-")
        c.drawString(220, y, r.get_status_display())
        c.drawString(300, y, f"{(r.distance_in_m or 0):.1f} m")
        y -= 16
    c.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="laporan_kehadiran.pdf")

@login_required
def daily_qr_page(request):
    if not request.user.is_staff:
        messages.error(request, "Menu QR harian hanya untuk pentadbir.")
        return redirect("dashboard")

    today = timezone.localdate()
    item, created = DailyQR.objects.get_or_create(
        date=today,
        defaults={"token": secrets.token_urlsafe(24), "active": True}
    )
    if request.method == "POST":
        item.token = secrets.token_urlsafe(24)
        item.active = True
        item.save()
        messages.success(request, "Kod QR harian baharu telah dijana.")

    import qrcode, base64
    qr_text = f"SKUA|{today.isoformat()}|{item.token}"
    img = qrcode.make(qr_text)
    stream = io.BytesIO()
    img.save(stream, format="PNG")
    qr_base64 = base64.b64encode(stream.getvalue()).decode("ascii")
    return render(request, "attendance/daily_qr.html", {
        "item": item,
        "qr_text": qr_text,
        "qr_base64": qr_base64,
    })


@login_required
def teacher_import_page(request):
    if not request.user.is_staff:
        messages.error(request, "Menu import guru hanya untuk pentadbir.")
        return redirect("dashboard")

    result = None
    if request.method == "POST":
        form = TeacherImportForm(request.POST, request.FILES)
        if form.is_valid():
            from openpyxl import load_workbook
            from django.contrib.auth import get_user_model
            from attendance.models import TeacherProfile

            workbook = load_workbook(form.cleaned_data["file"], data_only=True)
            sheet = workbook.active
            headers = {
                str(cell.value).strip().lower(): idx
                for idx, cell in enumerate(sheet[1])
                if cell.value
            }

            required = ["nama pengguna", "nama penuh", "emel", "no. staf", "jawatan"]
            missing = [h for h in required if h not in headers]
            if missing:
                messages.error(
                    request,
                    "Lajur tidak lengkap: " + ", ".join(missing)
                )
            else:
                User = get_user_model()
                created_count = 0
                updated_count = 0
                skipped = []

                for row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    username = str(row[headers["nama pengguna"]] or "").strip()
                    full_name = str(row[headers["nama penuh"]] or "").strip()
                    email = str(row[headers["emel"]] or "").strip()
                    staff_id = str(row[headers["no. staf"]] or "").strip()
                    position = str(row[headers["jawatan"]] or "").strip()

                    if not username or not full_name:
                        if any(v not in (None, "") for v in row):
                            skipped.append(row_no)
                        continue

                    user, created = User.objects.get_or_create(username=username)
                    names = full_name.split(maxsplit=1)
                    user.first_name = names[0]
                    user.last_name = names[1] if len(names) > 1 else ""
                    user.email = email
                    user.is_active = True
                    if created:
                        user.set_password(form.cleaned_data["default_password"])
                        created_count += 1
                    else:
                        updated_count += 1
                    user.save()

                    profile, _ = TeacherProfile.objects.get_or_create(user=user)
                    profile.staff_id = staff_id
                    profile.position = position
                    profile.save()

                result = {
                    "created": created_count,
                    "updated": updated_count,
                    "skipped": skipped,
                }
                messages.success(
                    request,
                    f"Import selesai: {created_count} akaun baharu, {updated_count} dikemas kini."
                )
    else:
        form = TeacherImportForm()

    return render(
        request,
        "attendance/teacher_import.html",
        {"form": form, "result": result},
    )


def health_check(request):
    from django.db import connection
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            cursor.fetchone()
        return JsonResponse({"status": "ok", "database": "connected", "version": "8.0"})
    except Exception:
        return JsonResponse(
            {"status": "error", "database": "unavailable", "version": "8.0"},
            status=503,
        )
